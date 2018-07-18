#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import tempfile
import copy

from openpyxl import Workbook, worksheet
from openpyxl.styles.borders import Border, Side

import excelreport

def workbook_create():
    wb = Workbook()
    for i in wb.worksheets:
        wb.remove(i)
    return wb

def sheet_create(wb, main_sheet_name):
    wb.create_sheet( main_sheet_name )
    ws = wb.active

    return ws

def sheet_print_setup(ws, print_settings):
    'https://openpyxl.readthedocs.io/en/2.5/_modules/openpyxl/worksheet/page.html'

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_options.headings = False
    ws.print_options.gridLines = False

    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False

    if print_settings in [ excelreport.PrintSet.LandscapeW1, excelreport.PrintSet.LandscapeW2 ]:
        worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='landscape')

    if print_settings in [ excelreport.PrintSet.PortraitW1 ]:
        worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation='portrait')

    if print_settings in [ excelreport.PrintSet.LandscapeW1, excelreport.PrintSet.PortraitW1 ]:
        ws.page_setup.fitToWidth = 1

    if print_settings in [ excelreport.PrintSet.LandscapeW2 ]:
        ws.print_options.horizontalCentered = False
        ws.page_setup.fitToWidth = 2

def temporary_file(filename='sample'):
    cnt = 1
    while True:
        newfilename = '{0:s}\\{1:s}{2:0>2d}.xlsx'.format( tempfile.gettempdir(), filename, cnt )
        if os.path.isfile(newfilename):
            cnt+=1
        else:
            break

    return newfilename

def apply_border( ws, start_row, start_column, end_row, end_column, border_style='thin' ):
    'https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/borders.html'

    thin_border = Border(left=Side(style=border_style),
                     right=Side(style=border_style),
                     top=Side(style=border_style),
                     bottom=Side(style=border_style))

    for i in range( start_row, end_row + 1 ):
        for j in range( start_column, end_column + 1 ):
            ws.cell( row=i, column=j ).border = thin_border

def apply_outline( ws, start_row, start_column, end_row, end_column, border_style='thin' ):
    ''
    side = Side(border_style=border_style)

    borderl = Border(left=side)
    borderr = Border(right=side)
    bordert = Border(top=side)
    borderb = Border(bottom=side)

    cells = [ ws.cell( row=i, column=start_column ) for i in range( start_row, end_row + 1 ) ]
    for cl in cells:
        border = copy.copy(cl.border)
        border.left = side
        cl.border = border

    cells = [ ws.cell( row=i, column=end_column ) for i in range( start_row, end_row + 1 ) ]
    for cl in cells:
        border = copy.copy(cl.border)
        border.right = side
        cl.border = border

    cells = [ ws.cell( row=start_row, column=i ) for i in range( start_column, end_column + 1 ) ]
    for cl in cells:
        border = copy.copy(cl.border)
        border.top = side
        cl.border = border

    cells = [ ws.cell( row=end_row, column=i ) for i in range( start_column, end_column + 1 ) ]
    for cl in cells:
        border = copy.copy(cl.border)
        border.bottom = side
        cl.border = border
