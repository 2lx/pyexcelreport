#!/usr/bin/python
# -*- coding: utf-8 -*-

from xlsutils import *
from xlsutils_apply import *
from xlscolor import *
from openpyxl.utils import get_column_letter

class XLSTableHeaderColumn:
    """Структура для хранения информации одного столбца (и подстолбцов) шапки таблицы
    """
    def __init__(self, title='', widths=[], struct=[]):
        self.title = title
        self.struct = struct
        self.widths = widths

        self.column_count = self._get_count()
        self.tree_height = self._get_height()

    def _get_count(self):
        """Подсчитывает количество колонок на листе у столбца
        """
        if self.struct: return sum(s.column_count for s in self.struct)
        elif self.widths: return len(self.widths)
        else: return 1

    def _get_height(self):
        """Подсчитывает высоту столбца в строках листа
        """
        return 1 + max([s.tree_height for s in self.struct], default=0)


class XLSTableHeader:
    """Класс, инкапсулирующий информацию и методы отображения шапки таблицы
    """
    def __init__(self, columns, bgcolor=Color.LT_GRAY.value, row_height=32, first_row_height=50):
        self._columns = columns
        self._bgcolor = bgcolor

        self.column_count = sum(cl.column_count for cl in columns)
        self.row_count = max([cl.tree_height for cl in columns], default = 0)
        self.default_row_height = row_height
        self.first_row_height = row_height

    def apply_widths(self, ws, first_col):
        """Применяет информацию о ширине столбцов из заголовка таблицы непосредственно к листу
        """
        def _traverse_leaves_and_set_width(col, cur_col):
            if col.struct:
                for icol in col.struct:
                    _traverse_leaves_and_set_width(icol, cur_col)
                    cur_col += icol.column_count
            else:
                for iwidth in col.widths:
                    ws.column_dimensions[get_column_letter(cur_col)].width = iwidth
                    cur_col += 1

        cur_col = first_col
        for col in self._columns:
            _traverse_leaves_and_set_width(col, cur_col)
            cur_col += col.column_count

    def apply(self, ws, first_row, first_col):
        """Отображает непосредственно в XLS шапку таблицы
        """
        def _traverse_tree_and_print_title(colinfo, cur_col, cur_height):
            start_row = first_row + cur_height
            end_row = start_row if colinfo.tree_height > 1 else first_row + self.row_count - 1
            end_col = cur_col + colinfo.column_count - 1

            apply_range(ws, start_row, cur_col, end_row, end_col, set_merge)
            ws.cell(row=start_row, column=cur_col).value = colinfo.title

            for coli in colinfo.struct:
                _traverse_tree_and_print_title(coli, cur_col, cur_height + 1)
                cur_col += coli.column_count


        if self.row_count == 1:
            ws.row_dimensions[first_row].height = self.first_row_height
        else:
            for i in range(0, self.row_count):
                ws.row_dimensions[first_row + i].height = self.default_row_height

        cur_col = first_col
        for colinfo in self._columns:
            _traverse_tree_and_print_title(colinfo, cur_col, 0)
            cur_col += colinfo.column_count

        clr = get_xlrange(first_row, first_col,
                          first_row + self.row_count - 1, cur_col - 1)
        apply_xlrange(ws, clr, set_borders)
        apply_xlrange(ws, clr, set_outline, border_style='medium')
        apply_xlrange(ws, clr, set_font, bold=True)
        apply_xlrange(ws, clr, set_alignment)
        apply_xlrange(ws, clr, set_fill, color=self._bgcolor)

        return first_row + self.row_count
