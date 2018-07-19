#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
from copy import copy

from openpyxl import Workbook, worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color, Alignment

def workbook_create():
    wb = Workbook()
    for i in wb.worksheets:
        wb.remove(i)
    return wb

def sheet_create(wb, main_sheet_name):
    wb.create_sheet( main_sheet_name )
    ws = wb.active

    return ws

def sheet_print_setup(ws, porientation, pwidth):
    'https://openpyxl.readthedocs.io/en/2.5/_modules/openpyxl/worksheet/page.html'

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_options.headings = False
    ws.print_options.gridLines = False

    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False

    worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation=porientation)
    ws.page_setup.fitToWidth = pwidth
    if pwidth == 2:
        ws.print_options.horizontalCentered = False

def most_bottom_right_coords(start_row, start_col, end_row, end_col):
    'Returns the most right and the most bottom coordinates'
    new_end_row = start_row if (end_row is None) or (end_row < start_row) else end_row
    new_end_col = start_col if (end_col is None) or (end_col < start_col) else end_col
    return new_end_row, new_end_col

def apply_border( ws, start_row, start_col, end_row=None, end_col=None, border_style='thin' ):
    'https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/borders.html'

    new_end_row, new_end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_border = Border(left=Side(style=border_style),
                     right=Side(style=border_style),
                     top=Side(style=border_style),
                     bottom=Side(style=border_style))

    for r in range( start_row, new_end_row + 1 ):
        for c in range( start_col, new_end_col + 1 ):
            ws.cell( row=r, column=c ).border = new_border

def apply_outline( ws, start_row, start_col, end_row=None, end_col=None, border_style='thin' ):
    ''
    def apply_border( cl, side_name ):
        border = copy(cl.border)
        getattr(border, side_name).border_style = border_style
        cl.border = border

    new_end_row, new_end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)

    for r in range( start_row, new_end_row + 1 ):
        apply_border(ws.cell(row=r, column=start_col), 'left')
        apply_border(ws.cell(row=r, column=end_col), 'right')

    for c in range( start_col, new_end_col + 1 ):
        apply_border(ws.cell(row=start_row, column=c), 'top')
        apply_border(ws.cell(row=end_row, column=c), 'bottom')

def font_setup(ws, start_row, start_col, end_row=None, end_col=None, \
            name='Calibri', size=11, bold=False, italic=False, underline='none', \
            vertAlign='baseline', strike=False, color='FF000000'):
    """https://openpyxl.readthedocs.io/en/2.5/styles.html
    """
    new_end_row, new_end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_font = Font(name=name, size=size, bold=bold, italic=italic, underline=underline, \
               vertAlign=vertAlign, strike=strike, color=color)

    for r in range(start_row, new_end_row + 1):
        for c in range(start_col, new_end_col + 1):
            ws.cell(row=r, column=c).font = new_font

def alignment_setup(ws, start_row, start_col, end_row=None, end_col=None, \
            horizontal='center', vertical='center', textRotation=None, wrapText=True, \
            shrinkToFit=True):
    """https://openpyxl.readthedocs.io/en/2.5/_modules/openpyxl/styles/alignment.html
    """
    new_end_row, new_end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_align = Alignment(horizontal=horizontal, vertical=vertical, textRotation=textRotation, \
                    wrapText=wrapText, shrinkToFit=shrinkToFit)

    for r in range(start_row, new_end_row + 1):
        for c in range(start_col, new_end_col + 1):
            ws.cell(row=r, column=c).alignment = new_align
