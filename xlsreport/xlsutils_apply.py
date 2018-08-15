#!/usr/bin/python
# -*- coding: utf-8 -*-

from copy import copy

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color, Alignment, PatternFill

def apply_cell(ws, start_row, start_col, f, **kwargs):
    f(ws, start_row, start_col, start_row, start_col, **kwargs)

def apply_range(ws, start_row, start_col, end_row, end_col, f, *args, **kwargs):
    if end_row < start_row: end_row, start_row = start_row, end_row
    if end_col < start_col: end_col, start_col = start_col, end_col
    f(ws, start_row, start_col, end_row, end_col, **kwargs)

def get_xlrange(start_row, start_col, end_row=None, end_col=None):
    if end_row is None: end_row = start_row
    if end_col is None: end_col = start_col
    if end_row < start_row:
        start_row, end_row = end_row, start_row
    if end_col < start_col:
        start_col, end_col = end_col, start_col

    return CellRange(min_row=start_row, min_col=start_col,
                     max_row=end_row,   max_col=end_col)

def apply_xlrange(ws, range, f, *args, **kwargs):
    f(ws, range.min_row, range.min_col, range.max_row, range.max_col, **kwargs)

"""Apply functions
"""

def set_merge(ws, start_row, start_col, end_row, end_col):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=end_row,     end_column=end_col)
    #  print("({0:d},{1:d}) - ({2:d},{3:d})".format(start_row, start_col, end_row, end_col))

def set_borders(ws, start_row, start_col, end_row, end_col, border_style='thin'):
    """
    """
    side = Side(style=border_style)
    new_border = Border(left=side, right=side, top=side, bottom=side)

    for r in range( start_row, end_row + 1 ):
        for c in range( start_col, end_col + 1 ):
            ws.cell( row=r, column=c ).border = new_border

def set_outline(ws, start_row, start_col, end_row, end_col, border_style='thin'):
    """
    """
    def _apply_border(cl, side_name):
        new_border = copy(cl.border)
        getattr(new_border, side_name).border_style = border_style
        cl.border = new_border

    for r in range(start_row, end_row + 1):
        _apply_border(ws.cell(row=r, column=start_col), 'left')
        _apply_border(ws.cell(row=r, column=end_col), 'right')

    for c in range(start_col, end_col + 1):
        _apply_border(ws.cell(row=start_row, column=c), 'top')
        _apply_border(ws.cell(row=end_row, column=c), 'bottom')

def set_font(ws, start_row, start_col, end_row, end_col,
             name='Calibri', size=11, bold=False, italic=False, underline='none',
             vertAlign='baseline', strike=False, color='FF000000'):
    """https://openpyxl.readthedocs.io/en/2.5/styles.html
    """
    new_font = Font(name=name, size=size, bold=bold, italic=italic, underline=underline,
                    vertAlign=vertAlign, strike=strike, color=color)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).font = new_font

def set_alignment(ws, start_row, start_col, end_row, end_col,
                  horizontal='center', vertical='center', textRotation=None, wrapText=True,
                  shrinkToFit=True):
    """https://openpyxl.readthedocs.io/en/2.5/_modules/openpyxl/styles/alignment.html
    """
    new_align = Alignment(horizontal=horizontal, vertical=vertical, textRotation=textRotation,
                    wrapText=wrapText, shrinkToFit=shrinkToFit)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).alignment = new_align

def set_fill(ws, start_row, start_col, end_row, end_col,
            color='FFFFFF', fill_type='solid'):
    """Fills the cell background with color
    """
    new_fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).fill = new_fill

def set_format(ws, start_row, start_col, end_row, end_col,
            format=''):
    """
    """
    if format == 'int':         new_format = '# ### ### ###'
    elif format == '1digit':  new_format   = '#,#0.0'
    elif format == 'currency':  new_format = '#,##0.00'
    elif format == '3digit':    new_format = '#,###0.000'
    else: return

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).number_format = new_format

