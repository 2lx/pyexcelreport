#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import sys
import datetime
from enum import Enum
from collections import namedtuple

from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter

from .xlslabel import *
from .xlstableheader import *
from .xlstable import *
from .xlsutils import *
from .systemutils import *
from .xlsutils_apply import *

PrintSetupStruct = namedtuple('PrintSetupStruct', 'orientation pages_width')

class PrintSetup(Enum):
    """Параметры печати печати
    """
    PortraitW1  = PrintSetupStruct('portrait',  1)
    LandscapeW1 = PrintSetupStruct('landscape', 1)
    LandscapeW2 = PrintSetupStruct('landscape', 2)

class XLSReport():
    """Класс, инкапсулирующий в себе методы для создания отчета в Excel
    """

    def __init__(self, sheet_name='Новый лист', print_setup=PrintSetup.LandscapeW1, protection=False):
        """Конструктор, создает книгу с одним именованным листом, устанавливает параметры для печати
        """
        self._wb = workbook_create()
        self._ws = sheet_create(self._wb, sheet_name)
        sheet_print_setup(self._ws, print_setup.value.orientation, print_setup.value.pages_width)
        self.protection = protection
        self._ws.protection.sheet = protection

    def append_sheet(self, sheet_name='Новый лист', print_setup=PrintSetup.LandscapeW1):
        """создает в конце книги еще один лист, устанавливает его параметры для печати
        """
        self._ws = sheet_create(self._wb, sheet_name)
        sheet_print_setup(self._ws, print_setup.value.orientation, print_setup.value.pages_width)
        self._ws.protection.sheet = self.protection

    def launch_excel(self, templatename='sample'):
        """Запускает программу по умолчанию для xls-файлов и открывает в ней workbook
        """
        newfilename = temporary_file(templatename)
        self._wb.save(newfilename)

        print("Открытие файла '{0:s}'...".format(newfilename))
        open_file(newfilename)

    def apply_column_widths(self, tableheader, first_col=1):
        tableheader.apply_widths(self._wb.active, first_col)

    def get_column_letter(self, col):
        return get_column_letter(col)

    def fold_columns_group(self, col1, col2):
        """Группирует колонки с возможностью скрытия, формат coli - latin uppercase
        """
        self._ws.column_dimensions.group(col1, col2, hidden=True)

    def print_preamble(self, max_col):
        self._ws.row_dimensions[1].height = 14
        self._ws.merge_cells(start_row=1, start_column=1,
                             end_row=1,   end_column=max_col)

        self._ws.cell(row=1, column=1).value = "Пользователь: {0:s}. Дата и время: {1:s}"\
                .format(os.getlogin(), datetime.datetime.now().strftime("%A %d %B %Y %H:%M"))
        apply_range(self._ws, 1, 1, 1, max_col, set_font, size=9, italic=True)
        apply_range(self._ws, 1, 1, 1, max_col, set_alignment, horizontal='right', vertical='top')
        return 2

    def print_label(self, label, first_row, first_col=1, col_count=1):
        label.apply(self._wb.active, first_row, first_col, col_count)
        return first_row + 1

    def print_tableheader(self, tableheader, first_row, first_col=1):
        return tableheader.apply(self._wb.active, first_row, first_col)

    def print_table(self, table, first_row, first_col=1):
        return table.apply(self._wb.active, first_row, first_col)
